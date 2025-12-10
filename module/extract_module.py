from io import BytesIO
from datetime import datetime

import openpyxl
from openpyxl import Workbook, load_workbook
from dateutil.relativedelta import relativedelta

from extract_fields import extract_all_fields


def calc_deadline(date_str, uu_tien):
    try:
        if not date_str or not uu_tien:
            return ""
        dt = datetime.strptime(date_str, "%m/%d/%Y")
        dt2 = dt + relativedelta(months=int(uu_tien))
        return dt2.strftime("%m/%d/%Y")
    except:
        return ""


def extract_only_kien_nghi(text: str):
    txt = text.lower()
    for key in ["đề nghị", "de nghi"]:
        pos = txt.find(key)
        if pos != -1:
            return text[pos:].strip()
    return text.strip()


def create_excel(kien_nghi_list, doi_tuong, so_van_ban, ngay_ban_hanh):
    wb = Workbook()
    ws = wb.active
    ws.title = "KPCS"

    columns = [
        "STT","Đối tượng được KT","Số văn bản",
        "Ngày, tháng, năm ban hành (mm/dd/yyyy)",
        "Tên Đoàn kiểm toán","Số hiệu rủi ro","Số hiệu kiểm soát",
        "Nghiệp vụ (R0)","Quy trình/hoạt động con (R1)","Tên phát hiện (R2)",
        "Chi tiết phát hiện (R3)","Dẫn chiếu","Mô tả chi tiết phát hiện",
        "CIF Khách hàng/bút toán","Tên khách hàng","Loại KH",
        "Số phát hiện/số mẫu chọn","Dư nợ sai phạm  (Triệu đồng)",
        "Số tiền tổn thất  (Triệu đồng)","Số tiền cần thu hồi (Triệu đồng)",
        "Trách nhiệm trực tiếp","Trách nhiệm quản lý",
        "Xếp hạng rủi ro","Xếp hạng kiểm soát",
        "Nguyên nhân","Ảnh hưởng","Kiến nghị",
        "Loại/nhóm nguyên nhân","Loại/nhóm ảnh hưởng",
        "Loại/nhóm kiến nghị","Chủ thể kiến nghị",
        "Kế hoạch thực hiện","Trách nhiệm thực hiện",
        "Đơn vị thực hiện KPCS","ĐVKD, AMC, Hội sở",
        "Người phê duyệt","Ý kiến của đơn vị",
        "Mức độ ưu tiên hành động",
        "Thời hạn hoàn thành (mm/dd/yyyy)",
        "Đã khắc phục","Ngày đã KPCS (mm/dd/yyyy)","CBKT"
    ]

    for idx, col in enumerate(columns, start=1):
        ws.cell(1, idx, col)

    col_idx = {name: i+1 for i, name in enumerate(columns)}

    for i, kn in enumerate(kien_nghi_list, start=2):
        ws.cell(i, col_idx["STT"], i - 1)
        ws.cell(i, col_idx["Đối tượng được KT"], doi_tuong or "")
        ws.cell(i, col_idx["Số văn bản"], so_van_ban or "")
        ws.cell(i, col_idx["Ngày, tháng, năm ban hành (mm/dd/yyyy)"], ngay_ban_hanh or "")

        # Chỉ lấy phần "Đề nghị..."
        only_kn = extract_only_kien_nghi(kn)
        ws.cell(i, col_idx["Kiến nghị"], only_kn)

        # Tách các trường bổ sung
        fields = extract_all_fields(kn)

        if fields["nguyen_nhan"]:
            ws.cell(i, col_idx["Nguyên nhân"], fields["nguyen_nhan"])

        if fields["uu_tien"]:
            ws.cell(i, col_idx["Mức độ ưu tiên hành động"], fields["uu_tien"])
            deadline = calc_deadline(ngay_ban_hanh, fields["uu_tien"])
            ws.cell(i, col_idx["Thời hạn hoàn thành (mm/dd/yyyy)"], deadline)

        if fields["nguoi_thuc_hien"]:
            ws.cell(i, col_idx["Trách nhiệm thực hiện"], fields["nguoi_thuc_hien"])

        if fields["nguoi_phe_duyet"]:
            ws.cell(i, col_idx["Người phê duyệt"], fields["nguoi_phe_duyet"])

        if fields["ngay_hoan_thanh"]:
            ws.cell(i, col_idx["Thời hạn hoàn thành (mm/dd/yyyy)"], fields["ngay_hoan_thanh"])

    out = BytesIO()
    wb.save(out)
    return out


def merge_kien_nghi(file_main, file_new):
    wb_main = load_workbook(file_main)
    ws_main = wb_main.active

    wb_new = load_workbook(file_new)
    ws_new = wb_new.active

    header = {ws_main.cell(1, c).value: c for c in range(1, ws_main.max_column+1)}

    def find_col(keys):
        for name, idx in header.items():
            if name and all(k in name.lower() for k in keys):
                return idx
        return None

    col_ngay_bh = find_col(["ngày", "ban hành"])
    col_uu = find_col(["mức độ ưu tiên"])
    col_dead = find_col(["thời hạn", "hoàn thành"])

    if col_dead is None:
        col_dead = ws_main.max_column + 1
        ws_main.cell(1, col_dead, "Thời hạn hoàn thành (mm/dd/yyyy)")

    for row in ws_new.iter_rows(min_row=2, values_only=True):
        new_r = ws_main.max_row + 1
        for c, v in enumerate(row, start=1):
            ws_main.cell(new_r, c, v)

        bh = ws_main.cell(new_r, col_ngay_bh).value if col_ngay_bh else None
        uu = ws_main.cell(new_r, col_uu).value if col_uu else None
        dl = calc_deadline(bh, uu)

        if dl:
            ws_main.cell(new_r, col_dead, dl)

    out = BytesIO()
    wb_main.save(out)
    return out

# from io import BytesIO
# from datetime import datetime

# import openpyxl
# from openpyxl import Workbook, load_workbook
# from dateutil.relativedelta import relativedelta

# from module.extract_fields import extract_all_fields


# def calc_deadline(date_str, uu_tien):
#     try:
#         if not date_str or not uu_tien:
#             return ""
#         dt = datetime.strptime(date_str, "%m/%d/%Y")
#         dt2 = dt + relativedelta(months=int(uu_tien))
#         return dt2.strftime("%m/%d/%Y")
#     except:
#         return ""


# def extract_only_kien_nghi(text: str):
#     txt = text.lower()
#     for key in ["đề nghị", "de nghi"]:
#         pos = txt.find(key)
#         if pos != -1:
#             return text[pos:].strip()
#     return text.strip()


# def create_excel(kien_nghi_list, doi_tuong, so_van_ban, ngay_ban_hanh):
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "KPCS"

#     columns = [
#         "STT","Đối tượng được KT","Số văn bản","Ngày ban hành",
#         "Tên Đoàn kiểm toán","Số hiệu rủi ro","Số hiệu kiểm soát",
#         "Nghiệp vụ (R0)","Quy trình (R1)","Tên phát hiện (R2)",
#         "Chi tiết phát hiện (R3)","Dẫn chiếu","Mô tả phát hiện",
#         "CIF","Tên khách hàng","Loại KH","Số phát hiện/mẫu chọn",
#         "Dư nợ sai phạm","Số tiền tổn thất","Số tiền cần thu hồi",
#         "Trách nhiệm trực tiếp","Trách nhiệm quản lý",
#         "Xếp hạng rủi ro","Xếp hạng kiểm soát",
#         "Nguyên nhân","Ảnh hưởng","Kiến nghị",
#         "Loại nguyên nhân","Loại ảnh hưởng","Loại kiến nghị",
#         "Chủ thể kiến nghị","Kế hoạch thực hiện",
#         "Trách nhiệm thực hiện","Đơn vị thực hiện KPCS",
#         "ĐVKD/AMC/Hội sở","Người phê duyệt","Ý kiến đơn vị",
#         "Mức độ ưu tiên","Thời hạn hoàn thành",
#         "Đã khắc phục","Ngày đã KPCS","CBKT"
#     ]

#     # Ghi header
#     for col_index, col_name in enumerate(columns, start=1):
#         ws.cell(1, col_index, col_name)

#     # Ghi dữ liệu
#     for i, kn in enumerate(kien_nghi_list, start=2):

#         ws.cell(i, 1, i - 1)                 # STT
#         ws.cell(i, 2, doi_tuong or "")       # Đối tượng KT
#         ws.cell(i, 3, so_van_ban or "")      # Số văn bản
#         ws.cell(i, 4, ngay_ban_hanh or "")   # Ngày ban hành
#         ws.cell(i, 27, kn)                   # Kiến nghị nội dung

#         # Các cột khác để trống
#         for col in range(5, len(columns) + 1):
#             if col != 27:
#                 ws.cell(i, col, "")

#     out = BytesIO()
#     wb.save(out)
#     return out



# def merge_kien_nghi(file_main, file_new):
#     wb_main = load_workbook(file_main)
#     ws_main = wb_main.active

#     wb_new = load_workbook(file_new)
#     ws_new = wb_new.active

#     header = {ws_main.cell(1, c).value: c for c in range(1, ws_main.max_column+1)}

#     def find_col(keys):
#         for name, idx in header.items():
#             if name and all(k in name.lower() for k in keys):
#                 return idx
#         return None

#     col_ngay_bh = find_col(["ngày", "ban hành"])
#     col_uu = find_col(["mức độ ưu tiên"])
#     col_dead = find_col(["thời hạn", "hoàn thành"])

#     if col_dead is None:
#         col_dead = ws_main.max_column + 1
#         ws_main.cell(1, col_dead, "Thời hạn hoàn thành (mm/dd/yyyy)")

#     for row in ws_new.iter_rows(min_row=2, values_only=True):
#         new_r = ws_main.max_row + 1
#         for c, v in enumerate(row, start=1):
#             ws_main.cell(new_r, c, v)

#         bh = ws_main.cell(new_r, col_ngay_bh).value if col_ngay_bh else None
#         uu = ws_main.cell(new_r, col_uu).value if col_uu else None
#         dl = calc_deadline(bh, uu)

#         if dl:
#             ws_main.cell(new_r, col_dead, dl)

#     out = BytesIO()
#     wb_main.save(out)
#     return out
