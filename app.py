import streamlit as st
import pytesseract
from PIL import Image
import numpy as np
import cv2
from pdf2image import convert_from_path
from docx import Document
from io import BytesIO
from PyPDF2 import PdfReader
import openpyxl
from openpyxl import Workbook, load_workbook


# ==== ĐỌC FILE WORD ====
def read_word(file):
    doc = Document(file)
    return "\n".join(p.text for p in doc.paragraphs)


# ==== ĐỌC PDF TEXT ====
def read_pdf(file):
    pdf = PdfReader(file)
    text = ""
    for page in pdf.pages:
        try:
            text += page.extract_text() + "\n"
        except:
            pass
    return text


# ==== OCR ẢNH ====
def ocr_image(img_file):
    img = Image.open(img_file).convert("RGB")
    img_np = np.array(img)

    gray = cv2.cvtColor(img_np, cv2.COLOR_BGR2GRAY)
    gray = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY)[1]

    text = pytesseract.image_to_string(gray, lang="vie")
    return text


# ==== OCR PDF SCAN ====
def ocr_pdf(file):
    pages = convert_from_path(file, dpi=250)
    result = ""

    for page in pages:
        img = np.array(page)
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        gray = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY)[1]

        text = pytesseract.image_to_string(gray, lang="vie")
        result += text + "\n"

    return result


# ==== TÁCH MỤC KIẾN NGHỊ ====
def extract_kien_nghi(text):
    text_lower = text.lower()

    start = text_lower.find("kiến nghị")
    if start == -1:
        return []

    section = text[start:]

    import re
    parts = re.split(r"\n\s*\d+[\.\)]\s+", section)

    results = []
    for p in parts:
        p = p.strip()
        if len(p) > 10 and not p.lower().startswith("kiến nghị"):
            results.append(p)

    return results


# ==== TÌM STT CUỐI TRONG FILE EXCEL GỐC ====
def get_last_stt_from_excel(file):
    try:
        wb = load_workbook(file)
        ws = wb.active
        last = 0
        for row in ws.iter_rows(min_row=2, max_col=1):
            cell = row[0].value
            if cell is not None and str(cell).isdigit():
                last = max(last, int(cell))
        return last
    except:
        return 0


# ==== TẠO FILE EXCEL MỚI + CỘNG DỒN STT ====
def create_excel(kien_nghi_list, last_stt):
    wb = Workbook()
    ws = wb.active
    ws.title = "KPCS"

    columns = [
        "STT","Đối tượng được KT","Số văn bản","Ngày ban hành",
        "Tên Đoàn kiểm toán","Số hiệu rủi ro","Số hiệu kiểm soát",
        "Nghiệp vụ (R0)","Quy trình (R1)","Tên phát hiện (R2)",
        "Chi tiết phát hiện (R3)","Dẫn chiếu","Mô tả phát hiện",
        "CIF","Tên khách hàng","Loại KH","Số phát hiện/mẫu chọn",
        "Dư nợ sai phạm","Số tiền tổn thất","Số tiền cần thu hồi",
        "Trách nhiệm trực tiếp","Trách nhiệm quản lý",
        "Xếp hạng rủi ro","Xếp hạng kiểm soát",
        "Nguyên nhân","Ảnh hưởng","Kiến nghị",
        "Loại nguyên nhân","Loại ảnh hưởng","Loại kiến nghị",
        "Chủ thể kiến nghị","Kế hoạch thực hiện",
        "Trách nhiệm thực hiện","Đơn vị thực hiện KPCS",
        "ĐVKD/AMC/Hội sở","Người phê duyệt","Ý kiến đơn vị",
        "Mức độ ưu tiên","Thời hạn hoàn thành",
        "Đã khắc phục","Ngày đã KPCS","CBKT"
    ]

    for col_index, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=col_index, value=col_name)

    for i, kn in enumerate(kien_nghi_list, start=2):
        ws.cell(i, 1, last_stt + (i - 1))
        ws.cell(i, 27, kn)

        for col in range(2, len(columns) + 1):
            if col != 27:
                ws.cell(i, col, "")

    output = BytesIO()
    wb.save(output)
    return output


# ================= STREAMLIT UI =================
st.title("📋 Công cụ Trích Kiến Nghị Báo Cáo Kiểm Toán – Full Version")
st.write("Hỗ trợ DOCX, PDF, PDF scan, ảnh; OCR tiếng Việt; tự động tạo Excel theo mẫu KPCS.")

uploaded = st.file_uploader("Upload báo cáo kiểm toán:", 
                             type=["pdf", "docx", "jpg", "jpeg", "png"])

excel_main = st.file_uploader("Upload file KPCS chính (để cộng dồn STT):", 
                              type=["xlsx"])


if uploaded:
    st.info("⏳ Đang đọc file...")

    ext = uploaded.name.split(".")[-1].lower()
    
    text = ""

    if ext in ["jpg", "png", "jpeg"]:
        text = ocr_image(uploaded)

    elif ext == "pdf":
        t = read_pdf(uploaded)
        if len(t.strip()) < 20:
            st.warning("PDF scan → dùng OCR...")
            text = ocr_pdf(uploaded)
        else:
            text = t

    elif ext == "docx":
        text = read_word(uploaded)

    st.subheader("📌 Preview văn bản")
    st.text_area("Dữ liệu OCR / Text:", text[:3000], height=200)

    kien_nghi = extract_kien_nghi(text)

    st.subheader(f"🔍 Tìm thấy {len(kien_nghi)} kiến nghị")

    if kien_nghi:

        last_stt = 0
        if excel_main:
            last_stt = get_last_stt_from_excel(excel_main)
            st.success(f"📌 STT cuối trong file chính: {last_stt}")

        excel_output = create_excel(kien_nghi, last_stt)

        st.download_button(
            label="⬇ Tải file Excel kiến nghị mới",
            data=excel_output.getvalue(),
            file_name="kien_nghi_moi.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
